import sys
import tkinter
import pandas as pd
from tkinter import ttk
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.label import DataLabel, DataLabelList
from tkinter import filedialog

dfmat = pd.read_csv('mat.csv')
dfexp = pd.read_csv('exp.csv',index_col=0)

root = tkinter.Tk()
root.title("データベース可視化ツール")
root.geometry("400x300")

Static1 = tkinter.Label(text=u'検索に含めるカラム')
Static1.pack()
Static1.place(x=0,y=40)
Static1 = tkinter.Label(text=u'---------------------------------------------------------------------')
Static1.pack()
Static1.place(x=0,y=80)
Static1 = tkinter.Label(text=u'サンプルID')
Static1.pack()
Static1.place(x=0,y=100)


EditBox = tkinter.Entry()
EditBox.insert(tkinter.END,"A")
EditBox.pack()
EditBox.place(x=50,y=0)

def search_value(event):
    value = EditBox.get()
    keywords = value.split()
    for i in range(len(names0)):
        if valscol[names0[i]].get():
            for keyword in keywords:
                for j in range(len(names)):
                    if keyword in dfmat.iloc[j,i]:
                        vals[names[j]].set(True)

def show_recip(event):
    r_root = tkinter.Toplevel(root)
    r_root.title("レシピ")   # ウィンドウタイトル
    r_root.geometry(str(80*len(list(dfmat.columns)))+"x200")
    tree = ttk.Treeview(r_root, columns=list(dfmat.columns))
    tree.column('#0',width=0, stretch='no')
    for col in dfmat.columns:
        tree.column(col,anchor='center',width=80)
        tree.heading(col, text=col,anchor='center')
    for i in range(len(dfmat)):
        if vals[names[i]].get():
            tree.insert(parent='', index='end', iid=i ,values=list(dfmat.iloc[i,:]))
    tree.pack()

def show_exp(event):
    plt.figure()
    x = dfexp.columns.astype(float)
    for i in range(len(names)):
        if vals[names[i]].get():
            plt.plot(x,dfexp.loc[names[i],:],label=names[i])
    plt.legend()
    plt.show(block=False)

def clear_col(event):
    for i in range(len(names0)):
        valscol[names0[i]].set(False)

def clear_ID(event):
    for i in range(len(names)):
        vals[names[i]].set(False)

def output_xlsx(event):
    # ワークブック作る
    wb = openpyxl.Workbook()
    # シート作る
    ws_mat = wb.worksheets[0]
    ws_mat.title = 'mat'
    ws_exp = wb.create_sheet(title="exp")
    # 数値書き込み(mat)
    for i in range(len(dfmat.columns)):
        ws_mat.cell(1,i+1).value = dfmat.columns[i]
    ind = 1
    for i in range(len(names)):
        if vals[names[i]].get():
            ls = list(dfmat.iloc[i,:])
            for j in range(len(ls)):
                ws_mat.cell(1+ind,j+1).value = ls[j]
            ind += 1
    # 数値書き込み(exp)
    ws_exp.cell(1,1).value = dfexp.index.name
    for i in range(len(dfexp.columns)):
        ws_exp.cell(1,i+2).value = dfexp.columns[i]
    ind = 1
    for i in range(len(names)):
        if vals[names[i]].get():
            ls = list(dfexp.iloc[i,:])
            ws_exp.cell(1+ind,1).value = names[i]
            for j in range(len(ls)):
                ws_exp.cell(1+ind,j+2).value = ls[j]
            ind += 1
    # 散布図
    c1 = ScatterChart()
    c1.width = 18     # デフォルト(15cm)
    c1.height = 10    # デフォルト(7cm)
    c1.title = "Scatter Chart"           # メインタイトル
    c1.x_axis.title = 'X'      # X軸のタイトル
    c1.y_axis.title = 'Y'    # Y軸のタイトル
    c1.legend.position = 'b'     # 凡例の配置位置
    for i in range(ind-1):
        values = Reference(ws_exp, min_row=i+2, max_row=i+2, min_col=1, max_col=len(dfexp.columns))
        xvalues = Reference(ws_exp, min_row=1, max_row=1, min_col=1, max_col=len(dfexp.columns))
        series = Series(values, xvalues, title_from_data=True)
        c1.series.append(series)
    ws_exp.add_chart(c1, "B13")
    
    filename = filedialog.asksaveasfilename(
        title = "名前を付けて保存",
        filetypes = [("xlsx", ".xlsx")], # ファイルフィルタ
        initialdir = "./", # 自分自身のディレクトリ
        defaultextension = "xlsx"
        )
    wb.save(filename)
    
Button1 = tkinter.Button(text=u'検索')
Button1.pack()
Button1.bind("<Button-1>",search_value) 
Button1.place(x=150,y=0)

Button2 = tkinter.Button(text=u'レシピ')
Button2.pack()
Button2.bind("<Button-1>",show_recip) 
Button2.place(x=200,y=0)

Button3 = tkinter.Button(text=u'実験結果')
Button3.pack()
Button3.bind("<Button-1>",show_exp) 
Button3.place(x=250,y=0)

Button3 = tkinter.Button(text=u'出力')
Button3.pack()
Button3.bind("<Button-1>",output_xlsx) 
Button3.place(x=320,y=0)

Button4 = tkinter.Button(text=u'クリア')
Button4.pack()
Button4.bind("<Button-1>",clear_col) 
Button4.place(x=100,y=37)

Button5 = tkinter.Button(text=u'クリア')
Button5.pack()
Button5.bind("<Button-1>",clear_ID) 
Button5.place(x=100,y=97)

names0 = list(dfmat.columns)
valscol = dict()
CheckBoxescol = dict()
for i in range(len(names0)):
    valscol[names0[i]] = tkinter.BooleanVar()
    valscol[names0[i]].set(False)
    CheckBoxescol[names0[i]]=tkinter.Checkbutton(text=names0[i], variable=valscol[names0[i]])
    CheckBoxescol[names0[i]].pack()
    CheckBoxescol[names0[i]].place(x=100*(i%4),y=60+20*(i//4))

names = list(dfmat.loc[:,'name'])
vals = dict()
CheckBoxes = dict()
for i in range(len(names)):
    vals[names[i]] = tkinter.BooleanVar()
    vals[names[i]].set(False)
    CheckBoxes[names[i]]=tkinter.Checkbutton(text=names[i], variable=vals[names[i]])
    CheckBoxes[names[i]].pack()
    CheckBoxes[names[i]].place(x=100*(i%4),y=120+20*(i//4))

plt.figure()
plt.close()

root.mainloop()
